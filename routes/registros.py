import io
from collections import OrderedDict
from datetime import date, datetime, timedelta, timezone
from zoneinfo import ZoneInfo
from flask import Blueprint, render_template, send_file, abort, request, redirect, url_for, flash
from sqlalchemy import func
from app import db
from models import Venta, SesionVenta

from routes.stand import get_stand_or_404

CHILE_TZ = ZoneInfo('America/Santiago')

registros_bp = Blueprint('registros', __name__, url_prefix='/s')

MESES = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
    5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
    9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
}


def parsear_fecha(fecha_str):
    """Parsea 'YYYY-MM-DD' y retorna date, o None si es invalido."""
    try:
        return datetime.strptime(fecha_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


def formato_fecha(fecha_str):
    """Convierte '2025-06-15' a '15 de Junio 2025'."""
    dt = parsear_fecha(fecha_str)
    if not dt:
        return fecha_str
    return f"{dt.day} de {MESES[dt.month]} {dt.year}"


def obtener_resumen_dia(stand, fecha_str):
    """Obtiene el resumen completo de un dia."""
    fecha = parsear_fecha(fecha_str)
    if not fecha:
        return None

    # Rango de timestamps para el dia (compatible con SQLite y PostgreSQL)
    fecha_inicio = datetime(fecha.year, fecha.month, fecha.day)
    fecha_fin = fecha_inicio + timedelta(days=1)

    ventas = Venta.query.filter(
        Venta.stand_id == stand.id,
        Venta.created_at >= fecha_inicio,
        Venta.created_at < fecha_fin
    ).order_by(Venta.created_at.asc()).all()

    if not ventas:
        return None

    total_ventas = len(ventas)
    total_recaudado = sum(v.total_final for v in ventas)
    total_pagado = sum(v.monto_pagado or 0 for v in ventas)
    total_pendiente_pago = total_recaudado - total_pagado

    ventas_efectivo = sum(1 for v in ventas if v.metodo_pago == 'efectivo')
    ventas_transferencia = sum(1 for v in ventas if v.metodo_pago == 'transferencia')
    monto_efectivo = sum(v.total_final for v in ventas if v.metodo_pago == 'efectivo')
    monto_transferencia = sum(v.total_final for v in ventas if v.metodo_pago == 'transferencia')

    pagadas = sum(1 for v in ventas if v.estado_pago == 'pagado')
    parciales = sum(1 for v in ventas if v.estado_pago == 'parcial')
    pago_pendiente = sum(1 for v in ventas if v.estado_pago == 'pendiente')

    entregadas = sum(1 for v in ventas if v.estado_entrega == 'entregado')
    listas = sum(1 for v in ventas if v.estado_entrega == 'listo')
    pendientes_entrega = sum(1 for v in ventas if v.estado_entrega == 'pendiente')

    # Productos mas vendidos del dia
    productos_vendidos = {}
    for v in ventas:
        for d in v.detalles:
            if d.nombre_producto not in productos_vendidos:
                productos_vendidos[d.nombre_producto] = {'cantidad': 0, 'total': 0}
            productos_vendidos[d.nombre_producto]['cantidad'] += d.cantidad
            productos_vendidos[d.nombre_producto]['total'] += d.subtotal

    productos_ranking = sorted(productos_vendidos.items(), key=lambda x: x[1]['cantidad'], reverse=True)

    return {
        'fecha': fecha_str,
        'fecha_formato': formato_fecha(fecha_str),
        'ventas': ventas,
        'total_ventas': total_ventas,
        'total_recaudado': total_recaudado,
        'total_pagado': total_pagado,
        'total_pendiente_pago': total_pendiente_pago,
        'ventas_efectivo': ventas_efectivo,
        'ventas_transferencia': ventas_transferencia,
        'monto_efectivo': monto_efectivo,
        'monto_transferencia': monto_transferencia,
        'pagadas': pagadas,
        'parciales': parciales,
        'pago_pendiente': pago_pendiente,
        'entregadas': entregadas,
        'listas': listas,
        'pendientes_entrega': pendientes_entrega,
        'productos_ranking': productos_ranking,
    }


@registros_bp.route('/<codigo>/registros')
def index(codigo):
    stand = get_stand_or_404(codigo)

    # Obtener todas las ventas y agrupar por dia en Python
    # (compatible con SQLite y PostgreSQL sin func.date())
    todas_ventas = stand.ventas.order_by(Venta.created_at.desc()).all()

    dias_dict = OrderedDict()
    for v in todas_ventas:
        local_dt = v.created_at.replace(tzinfo=timezone.utc).astimezone(CHILE_TZ)
        dia = local_dt.strftime('%Y-%m-%d')
        if dia not in dias_dict:
            dias_dict[dia] = {'total_ventas': 0, 'total_recaudado': 0, 'total_pagado': 0}
        dias_dict[dia]['total_ventas'] += 1
        dias_dict[dia]['total_recaudado'] += v.total_final
        dias_dict[dia]['total_pagado'] += v.monto_pagado or 0

    dias = []
    for fecha_str, data in dias_dict.items():
        dias.append({
            'fecha': fecha_str,
            'fecha_formato': formato_fecha(fecha_str),
            'total_ventas': data['total_ventas'],
            'total_recaudado': data['total_recaudado'],
            'total_pagado': data['total_pagado'],
            'pendiente': data['total_recaudado'] - data['total_pagado'],
        })

    total_general = sum(d['total_recaudado'] for d in dias)
    total_dias = len(dias)

    return render_template('stand/registros.html', stand=stand, dias=dias,
                           total_general=total_general, total_dias=total_dias)


@registros_bp.route('/<codigo>/registros/partial')
def index_partial(codigo):
    stand = get_stand_or_404(codigo)

    todas_ventas = stand.ventas.order_by(Venta.created_at.desc()).all()

    dias_dict = OrderedDict()
    for v in todas_ventas:
        local_dt = v.created_at.replace(tzinfo=timezone.utc).astimezone(CHILE_TZ)
        dia = local_dt.strftime('%Y-%m-%d')
        if dia not in dias_dict:
            dias_dict[dia] = {'total_ventas': 0, 'total_recaudado': 0, 'total_pagado': 0}
        dias_dict[dia]['total_ventas'] += 1
        dias_dict[dia]['total_recaudado'] += v.total_final
        dias_dict[dia]['total_pagado'] += v.monto_pagado or 0

    dias = []
    for fecha_str, data in dias_dict.items():
        dias.append({
            'fecha': fecha_str,
            'fecha_formato': formato_fecha(fecha_str),
            'total_ventas': data['total_ventas'],
            'total_recaudado': data['total_recaudado'],
            'total_pagado': data['total_pagado'],
            'pendiente': data['total_recaudado'] - data['total_pagado'],
        })

    total_general = sum(d['total_recaudado'] for d in dias)
    total_dias = len(dias)

    return render_template('stand/_registros_partial.html', stand=stand, dias=dias,
                           total_general=total_general, total_dias=total_dias)


@registros_bp.route('/<codigo>/registros/exportar-todo')
def exportar_todo_excel(codigo):
    stand = get_stand_or_404(codigo)

    ventas = Venta.query.filter_by(stand_id=stand.id).order_by(Venta.created_at.desc()).all()
    if not ventas:
        return 'No hay registros.', 404

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = 'Todas las Ventas'

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
    col_header_font = Font(name='Calibri', bold=True, size=10)
    col_header_fill = PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    day_fill = PatternFill(start_color='5856D6', end_color='5856D6', fill_type='solid')
    day_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)

    # Titulo
    ws.merge_cells('A1:I1')
    ws['A1'] = f'{stand.nombre} - Registro Completo de Ventas'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 30

    # Agrupar por dia
    ventas_por_dia = OrderedDict()
    for v in ventas:
        dia = v.created_at.strftime('%Y-%m-%d')
        if dia not in ventas_por_dia:
            ventas_por_dia[dia] = []
        ventas_por_dia[dia].append(v)

    row = 3
    total_global = 0

    for dia, ventas_dia in ventas_por_dia.items():
        total_dia = sum(v.total_final for v in ventas_dia)
        total_global += total_dia

        # Encabezado del dia
        ws.merge_cells(f'A{row}:I{row}')
        ws.cell(row=row, column=1, value=f'{formato_fecha(dia)} - {len(ventas_dia)} ventas - Total: ${total_dia:,}')
        ws.cell(row=row, column=1).font = day_font
        ws.cell(row=row, column=1).fill = day_fill
        ws.cell(row=row, column=1).alignment = center_align
        row += 1

        # Columnas
        cols = ['# Orden', 'Hora', 'Cliente', 'Producto', 'Cant.', 'P. Unit.', 'Subtotal', 'Total Venta', 'M. Pago']
        for i, col in enumerate(cols, 1):
            cell = ws.cell(row=row, column=i, value=col)
            cell.font = col_header_font
            cell.fill = col_header_fill
            cell.alignment = center_align
            cell.border = thin_border
        row += 1

        for venta in ventas_dia:
            for j, detalle in enumerate(venta.detalles):
                ws.cell(row=row, column=1, value=venta.numero_orden if j == 0 else '').alignment = center_align
                ws.cell(row=row, column=2, value=venta.created_at.strftime('%H:%M') if j == 0 else '').alignment = center_align
                ws.cell(row=row, column=3, value=(venta.cliente_nombre or 'S/N') if j == 0 else '').alignment = left_align
                ws.cell(row=row, column=4, value=detalle.nombre_producto).alignment = left_align
                ws.cell(row=row, column=5, value=detalle.cantidad).alignment = center_align
                ws.cell(row=row, column=6, value=detalle.precio_unitario).alignment = right_align
                ws.cell(row=row, column=6).number_format = '$#,##0'
                ws.cell(row=row, column=7, value=detalle.subtotal).alignment = right_align
                ws.cell(row=row, column=7).number_format = '$#,##0'
                ws.cell(row=row, column=8, value=venta.total_final if j == 0 else '').alignment = right_align
                if j == 0:
                    ws.cell(row=row, column=8).number_format = '$#,##0'
                ws.cell(row=row, column=9, value=venta.metodo_pago.capitalize() if j == 0 else '').alignment = center_align
                for c in range(1, 10):
                    ws.cell(row=row, column=c).border = thin_border
                row += 1

        row += 1

    # Total global
    ws.cell(row=row, column=7, value='TOTAL GLOBAL:').font = Font(bold=True, size=12)
    ws.cell(row=row, column=7).alignment = right_align
    ws.cell(row=row, column=8, value=total_global).font = Font(bold=True, size=12, color='2E7D32')
    ws.cell(row=row, column=8).number_format = '$#,##0'
    ws.cell(row=row, column=8).alignment = right_align

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 16

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{stand.nombre}_Registro_Completo.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@registros_bp.route('/<codigo>/registros/<fecha>')
def detalle_dia(codigo, fecha):
    stand = get_stand_or_404(codigo)

    # Validar formato de fecha
    if not parsear_fecha(fecha):
        abort(404)

    resumen = obtener_resumen_dia(stand, fecha)

    # Check if there's a session linked to this day's sales
    sesion_vinculada = None
    sesiones_disponibles = []
    if resumen:
        # Find session linked to any sale of this day
        for v in resumen['ventas']:
            if v.sesion_id:
                sesion_vinculada = SesionVenta.query.get(v.sesion_id)
                break
        # Get available sessions for linking
        sesiones_disponibles = stand.sesiones.order_by(SesionVenta.fecha.desc()).all()

    if not resumen:
        return render_template('stand/registro_dia.html', stand=stand, resumen=None, fecha=fecha)
    return render_template('stand/registro_dia.html', stand=stand, resumen=resumen, fecha=fecha,
                           sesion_vinculada=sesion_vinculada, sesiones_disponibles=sesiones_disponibles)


@registros_bp.route('/<codigo>/registros/<fecha>/excel')
def exportar_dia_excel(codigo, fecha):
    stand = get_stand_or_404(codigo)

    # Validar formato de fecha
    if not parsear_fecha(fecha):
        abort(404)

    resumen = obtener_resumen_dia(stand, fecha)
    if not resumen:
        return 'No hay registros para esta fecha.', 404

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # ===== Hoja 1: Detalle de Ventas =====
    ws = wb.active
    ws.title = 'Detalle Ventas'

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
    subheader_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    subheader_fill = PatternFill(start_color='5856D6', end_color='5856D6', fill_type='solid')
    col_header_font = Font(name='Calibri', bold=True, size=10)
    col_header_fill = PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid')
    money_font_green = Font(name='Calibri', bold=True, color='2E7D32')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    # Titulo
    ws.merge_cells('A1:H1')
    ws['A1'] = f'{stand.nombre} - Registro del {resumen["fecha_formato"]}'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 30

    # Resumen financiero
    ws.merge_cells('A3:H3')
    ws['A3'] = 'RESUMEN FINANCIERO'
    ws['A3'].font = subheader_font
    ws['A3'].fill = subheader_fill
    ws['A3'].alignment = center_align

    summary_data = [
        ('Total Ventas', resumen['total_ventas'], 'Total Recaudado', f"${resumen['total_recaudado']:,}"),
        ('Efectivo', f"{resumen['ventas_efectivo']} ventas (${resumen['monto_efectivo']:,})",
         'Transferencia', f"{resumen['ventas_transferencia']} ventas (${resumen['monto_transferencia']:,})"),
        ('Pagadas', resumen['pagadas'], 'Pendientes Pago', resumen['pago_pendiente']),
        ('Total Pagado', f"${resumen['total_pagado']:,}", 'Pendiente Cobro', f"${resumen['total_pendiente_pago']:,}"),
        ('Entregadas', resumen['entregadas'], 'Pendientes Entrega', resumen['pendientes_entrega']),
    ]

    row = 4
    for label1, val1, label2, val2 in summary_data:
        ws.cell(row=row, column=1, value=label1).font = Font(bold=True, size=10)
        ws.cell(row=row, column=2, value=str(val1)).alignment = left_align
        ws.cell(row=row, column=5, value=label2).font = Font(bold=True, size=10)
        ws.cell(row=row, column=6, value=str(val2)).alignment = left_align
        for c in range(1, 9):
            ws.cell(row=row, column=c).border = thin_border
        row += 1

    row += 1

    # Detalle de cada venta
    ws.merge_cells(f'A{row}:H{row}')
    ws.cell(row=row, column=1, value='DETALLE DE VENTAS').font = subheader_font
    ws.cell(row=row, column=1).fill = subheader_fill
    ws.cell(row=row, column=1).alignment = center_align
    row += 1

    # Encabezados de columnas
    cols = ['# Orden', 'Hora', 'Cliente', 'Producto', 'Cantidad', 'P. Unitario', 'Subtotal', 'M. Pago']
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=col)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1

    for venta in resumen['ventas']:
        for j, detalle in enumerate(venta.detalles):
            ws.cell(row=row, column=1, value=venta.numero_orden if j == 0 else '').alignment = center_align
            ws.cell(row=row, column=2, value=venta.created_at.strftime('%H:%M') if j == 0 else '').alignment = center_align
            ws.cell(row=row, column=3, value=(venta.cliente_nombre or 'S/N') if j == 0 else '').alignment = left_align
            ws.cell(row=row, column=4, value=detalle.nombre_producto).alignment = left_align
            ws.cell(row=row, column=5, value=detalle.cantidad).alignment = center_align
            ws.cell(row=row, column=6, value=detalle.precio_unitario).alignment = right_align
            ws.cell(row=row, column=6).number_format = '$#,##0'
            ws.cell(row=row, column=7, value=detalle.subtotal).alignment = right_align
            ws.cell(row=row, column=7).number_format = '$#,##0'
            ws.cell(row=row, column=8, value=venta.metodo_pago.capitalize() if j == 0 else '').alignment = center_align

            for c in range(1, 9):
                ws.cell(row=row, column=c).border = thin_border
            row += 1

        # Fila de total de la venta
        ws.cell(row=row, column=5, value='TOTAL:').font = Font(bold=True, size=10)
        ws.cell(row=row, column=5).alignment = right_align
        ws.cell(row=row, column=7, value=venta.total_final).font = money_font_green
        ws.cell(row=row, column=7).number_format = '$#,##0'
        ws.cell(row=row, column=7).alignment = right_align

        estado_pago_text = {'pagado': 'Pagado', 'parcial': 'Parcial', 'pendiente': 'Pendiente'}
        estado_entrega_text = {'entregado': 'Entregado', 'listo': 'Listo', 'pendiente': 'Pendiente'}
        ws.cell(row=row, column=3, value=f"Pago: {estado_pago_text.get(venta.estado_pago, venta.estado_pago)}").font = Font(italic=True, size=9)
        ws.cell(row=row, column=4, value=f"Entrega: {estado_entrega_text.get(venta.estado_entrega, venta.estado_entrega)}").font = Font(italic=True, size=9)

        if venta.notas:
            ws.cell(row=row, column=1, value=f"Nota: {venta.notas}").font = Font(italic=True, size=9, color='666666')

        for c in range(1, 9):
            ws.cell(row=row, column=c).border = thin_border
        row += 1

    # Fila total general
    row += 1
    ws.cell(row=row, column=5, value='TOTAL GENERAL:').font = Font(bold=True, size=12)
    ws.cell(row=row, column=5).alignment = right_align
    ws.cell(row=row, column=7, value=resumen['total_recaudado']).font = Font(bold=True, size=12, color='2E7D32')
    ws.cell(row=row, column=7).number_format = '$#,##0'
    ws.cell(row=row, column=7).alignment = right_align

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 16

    # ===== Hoja 2: Productos del Dia =====
    ws2 = wb.create_sheet('Productos del Dia')

    ws2.merge_cells('A1:D1')
    ws2['A1'] = f'Productos Vendidos - {resumen["fecha_formato"]}'
    ws2['A1'].font = header_font
    ws2['A1'].fill = header_fill
    ws2['A1'].alignment = center_align
    ws2.row_dimensions[1].height = 30

    prod_cols = ['Producto', 'Cantidad Vendida', 'Total Vendido', '% del Total']
    for i, col in enumerate(prod_cols, 1):
        cell = ws2.cell(row=3, column=i, value=col)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = center_align
        cell.border = thin_border

    row2 = 4
    for nombre, datos in resumen['productos_ranking']:
        porcentaje = (datos['total'] / resumen['total_recaudado'] * 100) if resumen['total_recaudado'] > 0 else 0
        ws2.cell(row=row2, column=1, value=nombre).alignment = left_align
        ws2.cell(row=row2, column=2, value=datos['cantidad']).alignment = center_align
        ws2.cell(row=row2, column=3, value=datos['total']).alignment = right_align
        ws2.cell(row=row2, column=3).number_format = '$#,##0'
        ws2.cell(row=row2, column=4, value=round(porcentaje, 1)).alignment = center_align
        ws2.cell(row=row2, column=4).number_format = '0.0"%"'
        for c in range(1, 5):
            ws2.cell(row=row2, column=c).border = thin_border
        row2 += 1

    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 16
    ws2.column_dimensions['D'].width = 14

    # Guardar a memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{stand.nombre}_Registro_{fecha}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@registros_bp.route('/<codigo>/registros/<fecha>/vincular-sesion', methods=['POST'])
def vincular_sesion(codigo, fecha):
    stand = get_stand_or_404(codigo)

    fecha_date = parsear_fecha(fecha)
    if not fecha_date:
        abort(404)

    sesion_id = request.form.get('sesion_id', type=int)
    crear_nueva = request.form.get('crear_nueva')

    if crear_nueva:
        nombre = request.form.get('nombre', '').strip()
        sesion = SesionVenta(
            stand_id=stand.id,
            fecha=fecha_date,
            nombre=nombre or None,
            estado='abierta'
        )
        db.session.add(sesion)
        db.session.flush()
        sesion_id = sesion.id
    elif sesion_id:
        sesion = SesionVenta.query.filter_by(id=sesion_id, stand_id=stand.id).first()
        if not sesion:
            flash('Sesion no encontrada.', 'danger')
            return redirect(url_for('registros.detalle_dia', codigo=codigo, fecha=fecha))
    else:
        flash('Selecciona una sesion o crea una nueva.', 'danger')
        return redirect(url_for('registros.detalle_dia', codigo=codigo, fecha=fecha))

    # Update all sales of this day to link to the session
    fecha_inicio = datetime(fecha_date.year, fecha_date.month, fecha_date.day)
    fecha_fin = fecha_inicio + timedelta(days=1)

    ventas = Venta.query.filter(
        Venta.stand_id == stand.id,
        Venta.created_at >= fecha_inicio,
        Venta.created_at < fecha_fin
    ).all()

    for v in ventas:
        v.sesion_id = sesion_id

    db.session.commit()
    flash(f'Sesion vinculada a {len(ventas)} ventas.', 'success')
    return redirect(url_for('registros.detalle_dia', codigo=codigo, fecha=fecha))
