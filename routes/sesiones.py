import io
from datetime import date, datetime, timezone
from zoneinfo import ZoneInfo
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from sqlalchemy import func
from app import db
from models import SesionVenta, SesionIntegrante, Integrante, Venta
from routes.stand import get_stand_or_404

CHILE_TZ = ZoneInfo('America/Santiago')

MESES = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
    5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
    9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
}

sesiones_bp = Blueprint('sesiones', __name__, url_prefix='/s')


def formato_fecha_sesion(fecha):
    """Convierte date a '15 de Junio 2025'."""
    return f"{fecha.day} de {MESES[fecha.month]} {fecha.year}"


def obtener_resumen_sesion(ventas):
    """Calcula resumen financiero completo para una lista de ventas."""
    total_ventas = len(ventas)
    total_recaudado = sum(v.total_final for v in ventas)
    total_pagado = sum(v.monto_pagado or 0 for v in ventas)
    total_pendiente_pago = total_recaudado - total_pagado

    ventas_efectivo = sum(1 for v in ventas if v.metodo_pago == 'efectivo')
    ventas_transferencia = sum(1 for v in ventas if v.metodo_pago == 'transferencia')
    monto_efectivo = sum(v.total_final for v in ventas if v.metodo_pago == 'efectivo')
    monto_transferencia = sum(v.total_final for v in ventas if v.metodo_pago == 'transferencia')

    pagadas = sum(1 for v in ventas if v.estado_pago == 'pagado')
    parciales = sum(1 for v in ventas if v.estado_pago == 'parcial')
    pago_pendiente = sum(1 for v in ventas if v.estado_pago == 'pendiente')

    entregadas = sum(1 for v in ventas if v.estado_entrega == 'entregado')
    listas = sum(1 for v in ventas if v.estado_entrega == 'listo')
    pendientes_entrega = sum(1 for v in ventas if v.estado_entrega == 'pendiente')

    productos_vendidos = {}
    for v in ventas:
        for d in v.detalles:
            if d.nombre_producto not in productos_vendidos:
                productos_vendidos[d.nombre_producto] = {'cantidad': 0, 'total': 0}
            productos_vendidos[d.nombre_producto]['cantidad'] += d.cantidad
            productos_vendidos[d.nombre_producto]['total'] += d.subtotal

    productos_ranking = sorted(productos_vendidos.items(), key=lambda x: x[1]['cantidad'], reverse=True)

    return {
        'ventas': ventas,
        'total_ventas': total_ventas,
        'total_recaudado': total_recaudado,
        'total_pagado': total_pagado,
        'total_pendiente_pago': total_pendiente_pago,
        'ventas_efectivo': ventas_efectivo,
        'ventas_transferencia': ventas_transferencia,
        'monto_efectivo': monto_efectivo,
        'monto_transferencia': monto_transferencia,
        'pagadas': pagadas,
        'parciales': parciales,
        'pago_pendiente': pago_pendiente,
        'entregadas': entregadas,
        'listas': listas,
        'pendientes_entrega': pendientes_entrega,
        'productos_ranking': productos_ranking,
    }


@sesiones_bp.route('/<codigo>/sesiones')
def lista(codigo):
    stand = get_stand_or_404(codigo)
    sesiones = stand.sesiones.order_by(SesionVenta.fecha.desc()).all()

    sesiones_data = []
    total_general = 0
    for s in sesiones:
        ventas = s.ventas.all()
        total_recaudado = sum(v.total_final for v in ventas)
        total_pagado = sum(v.monto_pagado or 0 for v in ventas)
        pendiente = total_recaudado - total_pagado
        total_general += total_recaudado
        sesiones_data.append({
            'sesion': s,
            'total_ventas': len(ventas),
            'total_recaudado': total_recaudado,
            'total_pagado': total_pagado,
            'pendiente': pendiente,
            'equipo': len(s.integrantes),
            'fecha_formato': formato_fecha_sesion(s.fecha),
        })

    return render_template('sesiones/lista.html', stand=stand,
                           sesiones=sesiones_data, total_general=total_general,
                           total_sesiones=len(sesiones_data), today=date.today().isoformat())


@sesiones_bp.route('/<codigo>/sesiones/nueva', methods=['POST'])
def nueva(codigo):
    stand = get_stand_or_404(codigo)
    fecha_str = request.form.get('fecha', '')
    nombre = request.form.get('nombre', '').strip()

    try:
        fecha = date.fromisoformat(fecha_str) if fecha_str else date.today()
    except ValueError:
        fecha = date.today()

    sesion = SesionVenta(
        stand_id=stand.id,
        fecha=fecha,
        nombre=nombre or None,
        estado='programada'
    )
    db.session.add(sesion)
    db.session.commit()
    flash('Sesion creada.', 'success')
    return redirect(url_for('sesiones.detalle', codigo=codigo, sesion_id=sesion.id))


@sesiones_bp.route('/<codigo>/sesiones/<int:sesion_id>')
def detalle(codigo, sesion_id):
    stand = get_stand_or_404(codigo)
    sesion = SesionVenta.query.filter_by(id=sesion_id, stand_id=stand.id).first_or_404()
    ventas = sesion.ventas.order_by(Venta.created_at.desc()).all()
    integrantes_disponibles = stand.integrantes.filter_by(activo=True).order_by(Integrante.nombre).all()

    resumen = obtener_resumen_sesion(ventas)

    return render_template('sesiones/detalle.html', stand=stand, sesion=sesion,
                           resumen=resumen, integrantes_disponibles=integrantes_disponibles)


@sesiones_bp.route('/<codigo>/sesiones/<int:sesion_id>/estado', methods=['POST'])
def cambiar_estado(codigo, sesion_id):
    stand = get_stand_or_404(codigo)
    sesion = SesionVenta.query.filter_by(id=sesion_id, stand_id=stand.id).first_or_404()
    nuevo_estado = request.form.get('estado')

    if nuevo_estado in ('programada', 'abierta', 'cerrada'):
        sesion.estado = nuevo_estado
        db.session.commit()
        flash(f'Sesion {nuevo_estado}.', 'success')

    return redirect(url_for('sesiones.detalle', codigo=codigo, sesion_id=sesion.id))


@sesiones_bp.route('/<codigo>/sesiones/<int:sesion_id>/nuevo-integrante', methods=['POST'])
def nuevo_integrante(codigo, sesion_id):
    stand = get_stand_or_404(codigo)
    sesion = SesionVenta.query.filter_by(id=sesion_id, stand_id=stand.id).first_or_404()

    nombre = request.form.get('nombre', '').strip()
    telefono = request.form.get('telefono', '').strip()

    if not nombre:
        flash('El nombre es obligatorio.', 'danger')
        return redirect(url_for('sesiones.detalle', codigo=codigo, sesion_id=sesion.id))

    integrante = Integrante(
        stand_id=stand.id,
        nombre=nombre,
        telefono=telefono or None
    )
    db.session.add(integrante)
    db.session.commit()
    flash(f'Integrante "{nombre}" creado.', 'success')
    return redirect(url_for('sesiones.detalle', codigo=codigo, sesion_id=sesion.id))


@sesiones_bp.route('/<codigo>/sesiones/<int:sesion_id>/integrantes', methods=['POST'])
def gestionar_integrantes(codigo, sesion_id):
    stand = get_stand_or_404(codigo)
    sesion = SesionVenta.query.filter_by(id=sesion_id, stand_id=stand.id).first_or_404()

    # Remove existing assignments
    SesionIntegrante.query.filter_by(sesion_id=sesion.id).delete()

    # Add new assignments from form
    integrantes_disponibles = stand.integrantes.filter_by(activo=True).all()
    for integrante in integrantes_disponibles:
        roles = request.form.getlist(f'roles_{integrante.id}')
        for rol in roles:
            if rol in ('cocina', 'atencion', 'entrega'):
                si = SesionIntegrante(
                    sesion_id=sesion.id,
                    integrante_id=integrante.id,
                    rol=rol
                )
                db.session.add(si)

    db.session.commit()
    flash('Roles actualizados.', 'success')
    return redirect(url_for('sesiones.detalle', codigo=codigo, sesion_id=sesion.id))


@sesiones_bp.route('/<codigo>/sesiones/<int:sesion_id>/excel')
def exportar_sesion_excel(codigo, sesion_id):
    stand = get_stand_or_404(codigo)
    sesion = SesionVenta.query.filter_by(id=sesion_id, stand_id=stand.id).first_or_404()
    ventas = sesion.ventas.order_by(Venta.created_at.asc()).all()

    if not ventas:
        return 'No hay ventas en esta sesion.', 404

    resumen = obtener_resumen_sesion(ventas)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = 'Detalle Ventas'

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
    subheader_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    subheader_fill = PatternFill(start_color='5856D6', end_color='5856D6', fill_type='solid')
    col_header_font = Font(name='Calibri', bold=True, size=10)
    col_header_fill = PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid')
    money_font_green = Font(name='Calibri', bold=True, color='2E7D32')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    titulo = f'{stand.nombre} - Sesion {formato_fecha_sesion(sesion.fecha)}'
    if sesion.nombre:
        titulo += f' ({sesion.nombre})'

    ws.merge_cells('A1:H1')
    ws['A1'] = titulo
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 30

    # Resumen financiero
    ws.merge_cells('A3:H3')
    ws['A3'] = 'RESUMEN FINANCIERO'
    ws['A3'].font = subheader_font
    ws['A3'].fill = subheader_fill
    ws['A3'].alignment = center_align

    summary_data = [
        ('Total Ventas', resumen['total_ventas'], 'Total Recaudado', f"${resumen['total_recaudado']:,}"),
        ('Efectivo', f"{resumen['ventas_efectivo']} ventas (${resumen['monto_efectivo']:,})",
         'Transferencia', f"{resumen['ventas_transferencia']} ventas (${resumen['monto_transferencia']:,})"),
        ('Pagadas', resumen['pagadas'], 'Pendientes Pago', resumen['pago_pendiente']),
        ('Total Pagado', f"${resumen['total_pagado']:,}", 'Pendiente Cobro', f"${resumen['total_pendiente_pago']:,}"),
        ('Entregadas', resumen['entregadas'], 'Pendientes Entrega', resumen['pendientes_entrega']),
    ]

    row = 4
    for label1, val1, label2, val2 in summary_data:
        ws.cell(row=row, column=1, value=label1).font = Font(bold=True, size=10)
        ws.cell(row=row, column=2, value=str(val1)).alignment = left_align
        ws.cell(row=row, column=5, value=label2).font = Font(bold=True, size=10)
        ws.cell(row=row, column=6, value=str(val2)).alignment = left_align
        for c in range(1, 9):
            ws.cell(row=row, column=c).border = thin_border
        row += 1

    row += 1

    # Detalle de ventas
    ws.merge_cells(f'A{row}:H{row}')
    ws.cell(row=row, column=1, value='DETALLE DE VENTAS').font = subheader_font
    ws.cell(row=row, column=1).fill = subheader_fill
    ws.cell(row=row, column=1).alignment = center_align
    row += 1

    cols = ['# Orden', 'Hora', 'Cliente', 'Producto', 'Cantidad', 'P. Unitario', 'Subtotal', 'M. Pago']
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=col)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1

    for venta in ventas:
        hora_chile = venta.created_at.replace(tzinfo=timezone.utc).astimezone(CHILE_TZ).strftime('%H:%M')
        for j, detalle in enumerate(venta.detalles):
            ws.cell(row=row, column=1, value=venta.numero_orden if j == 0 else '').alignment = center_align
            ws.cell(row=row, column=2, value=hora_chile if j == 0 else '').alignment = center_align
            ws.cell(row=row, column=3, value=(venta.cliente_nombre or 'S/N') if j == 0 else '').alignment = left_align
            ws.cell(row=row, column=4, value=detalle.nombre_producto).alignment = left_align
            ws.cell(row=row, column=5, value=detalle.cantidad).alignment = center_align
            ws.cell(row=row, column=6, value=detalle.precio_unitario).alignment = right_align
            ws.cell(row=row, column=6).number_format = '$#,##0'
            ws.cell(row=row, column=7, value=detalle.subtotal).alignment = right_align
            ws.cell(row=row, column=7).number_format = '$#,##0'
            ws.cell(row=row, column=8, value=venta.metodo_pago.capitalize() if j == 0 else '').alignment = center_align

            for c in range(1, 9):
                ws.cell(row=row, column=c).border = thin_border
            row += 1

        # Total row
        ws.cell(row=row, column=5, value='TOTAL:').font = Font(bold=True, size=10)
        ws.cell(row=row, column=5).alignment = right_align
        ws.cell(row=row, column=7, value=venta.total_final).font = money_font_green
        ws.cell(row=row, column=7).number_format = '$#,##0'
        ws.cell(row=row, column=7).alignment = right_align

        estado_pago_text = {'pagado': 'Pagado', 'parcial': 'Parcial', 'pendiente': 'Pendiente'}
        estado_entrega_text = {'entregado': 'Entregado', 'listo': 'Listo', 'pendiente': 'Pendiente'}
        ws.cell(row=row, column=3, value=f"Pago: {estado_pago_text.get(venta.estado_pago, venta.estado_pago)}").font = Font(italic=True, size=9)
        ws.cell(row=row, column=4, value=f"Entrega: {estado_entrega_text.get(venta.estado_entrega, venta.estado_entrega)}").font = Font(italic=True, size=9)

        if venta.notas:
            ws.cell(row=row, column=1, value=f"Nota: {venta.notas}").font = Font(italic=True, size=9, color='666666')

        for c in range(1, 9):
            ws.cell(row=row, column=c).border = thin_border
        row += 1

    # Total general
    row += 1
    ws.cell(row=row, column=5, value='TOTAL GENERAL:').font = Font(bold=True, size=12)
    ws.cell(row=row, column=5).alignment = right_align
    ws.cell(row=row, column=7, value=resumen['total_recaudado']).font = Font(bold=True, size=12, color='2E7D32')
    ws.cell(row=row, column=7).number_format = '$#,##0'
    ws.cell(row=row, column=7).alignment = right_align

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 16

    # Hoja 2: Productos
    ws2 = wb.create_sheet('Productos')

    ws2.merge_cells('A1:D1')
    ws2['A1'] = f'Productos Vendidos - Sesion {formato_fecha_sesion(sesion.fecha)}'
    ws2['A1'].font = header_font
    ws2['A1'].fill = header_fill
    ws2['A1'].alignment = center_align
    ws2.row_dimensions[1].height = 30

    prod_cols = ['Producto', 'Cantidad Vendida', 'Total Vendido', '% del Total']
    for i, col in enumerate(prod_cols, 1):
        cell = ws2.cell(row=3, column=i, value=col)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = center_align
        cell.border = thin_border

    row2 = 4
    for nombre, datos in resumen['productos_ranking']:
        porcentaje = (datos['total'] / resumen['total_recaudado'] * 100) if resumen['total_recaudado'] > 0 else 0
        ws2.cell(row=row2, column=1, value=nombre).alignment = left_align
        ws2.cell(row=row2, column=2, value=datos['cantidad']).alignment = center_align
        ws2.cell(row=row2, column=3, value=datos['total']).alignment = right_align
        ws2.cell(row=row2, column=3).number_format = '$#,##0'
        ws2.cell(row=row2, column=4, value=round(porcentaje, 1)).alignment = center_align
        ws2.cell(row=row2, column=4).number_format = '0.0"%"'
        for c in range(1, 5):
            ws2.cell(row=row2, column=c).border = thin_border
        row2 += 1

    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 16
    ws2.column_dimensions['D'].width = 14

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fecha_str = sesion.fecha.strftime('%Y-%m-%d')
    filename = f"{stand.nombre}_Sesion_{fecha_str}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)
